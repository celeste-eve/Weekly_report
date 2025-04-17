import pandas as pd
import sys
import os
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
import datetime
from fpdf import FPDF
import matplotlib.pyplot as plt

# Get current timestamp
D = datetime.datetime.now().strftime("%Y-%m-%d %H_%M")

def generate_summary(input_file, output_dir):
    # Load sheet with largest numeric name
    xls = pd.ExcelFile(input_file)
    numeric_sheets = [int(name) for name in xls.sheet_names if name.isdigit()]
    
    if not numeric_sheets:
        print("No numeric sheet names found.")
        return
    
    sheet_name = str(max(numeric_sheets))
    df = pd.read_excel(input_file, sheet_name=sheet_name, header=1)

    # Rename columns
    df.columns.values[0] = 'Project'
    df.columns.values[1] = 'Test Type'

    columns_to_keep = ['Project', 'Broad type', 'Test Type', 'Scheduled', 'Checked', 'InProgress', 'Completed', 'Approved', 'Total', 'NCF? Restricted numbers', 'Week Approved Increase']
    df = df[[col for col in columns_to_keep if col in df.columns]]

    # Cleanup and calculations
    df['NCF? Restricted numbers'] = pd.to_numeric(df['NCF? Restricted numbers'], errors='coerce').abs()
    df['Week Approved Increase'] = pd.to_numeric(df['Week Approved Increase'], errors='coerce').clip(lower=0)

    first_col = 'Project'
    type_col = 'Broad type'

    grouped = df.groupby([first_col])
    summary_df = grouped.agg({
        'Completed': 'sum',
        'Approved': 'sum',
        'Checked': 'sum',
        'Scheduled': 'sum',
        'InProgress': 'sum',
        'NCF? Restricted numbers': 'sum',
        'Week Approved Increase': 'sum'
    }).reset_index()

    summary_df['Total Completed'] = summary_df['Completed'] + summary_df['Approved'] + summary_df['Checked']
    summary_df['Total Remaining'] = summary_df['Scheduled'] + summary_df['InProgress']

    final_df = summary_df[[first_col, 'Total Completed', 'Total Remaining', 'NCF? Restricted numbers', 'Week Approved Increase']]
    final_df = final_df.rename(columns={'NCF? Restricted numbers': 'Total NCF? Restricted numbers'})

    grouped = df.groupby([type_col, first_col])
    grouped_df = grouped.agg({
        'Completed': 'sum',
        'Approved': 'sum',
        'Checked': 'sum',
        'Scheduled': 'sum',
        'InProgress': 'sum',
        'NCF? Restricted numbers': 'sum',
        'Week Approved Increase': 'sum'
    }).reset_index()

    grouped_df['Total Completed'] = grouped_df['Completed'] + grouped_df['Approved'] + grouped_df['Checked']
    grouped_df['Total Remaining'] = grouped_df['Scheduled'] + grouped_df['InProgress']

    groupedC_df = grouped_df[[type_col, first_col, 'Total Completed', 'Total Remaining', 'NCF? Restricted numbers', 'Week Approved Increase']]
    groupedC_df = groupedC_df.rename(columns={'NCF? Restricted numbers': 'Total NCF? Restricted numbers'})

    output_file = os.path.join(output_dir, f"summary_output_{D}.xlsx")

    # CREATING A PDF REPORT
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt="Weekly Summary Report", ln=True, align='C')

    def get_column_widths(df, min_width=20, max_width=45):
        # Estimate column widths from content
        col_widths = []
        for col in df.columns:
            max_content = max([len(str(val)) for val in df[col]] + [len(str(col))])
            width = min(max(len(str(col)) * 2, max_content * 2), max_width)
            col_widths.append(max(min_width, width))
        return col_widths

    def add_table_to_pdf(dataframe, title):
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(200, 10, txt=title.encode('latin-1', 'replace').decode('latin-1'), ln=True, align='C')
        pdf.set_font("Arial", size=8)

        col_widths = get_column_widths(dataframe) # variable column widths
        line_height = 5

        # Header
        pdf.set_font("Arial", 'B', 8) 
        y_before = pdf.get_y()
        x_start = pdf.get_x()
        for i, col in enumerate(dataframe.columns):
            pdf.set_fill_color(200, 200, 200)  # Light gray fill
            header = str(col).encode('latin-1', 'replace').decode('latin-1')
            pdf.multi_cell(col_widths[i], line_height, header, border=1, align='C')
            pdf.set_xy(x_start + sum(col_widths[:i+1]), y_before)
        pdf.ln(line_height * 2)

        # Data Rows
        pdf.set_font("Arial", '', 8)
        for _, row in dataframe.iterrows():
            x_start = pdf.get_x()
            y_start = pdf.get_y()
            max_y = y_start
            heights = []

            # First pass to calculate max height in this row
            for i, item in enumerate(row):
                text = str(item).encode('latin-1', 'replace').decode('latin-1')
                lines = pdf.multi_cell(col_widths[i], line_height, text, border=0, align='L', split_only=True)
                height = len(lines) * line_height
                heights.append(height)
                max_y = max(max_y, y_start + height)

            # Second pass to actually print the cells
            for i, item in enumerate(row):
                text = str(item).encode('latin-1', 'replace').decode('latin-1')
                pdf.set_xy(x_start + sum(col_widths[:i]), y_start)
                pdf.multi_cell(col_widths[i], line_height, text, border=1, align='L')

            pdf.set_y(max_y)

    # PDF Graphs
    def add_figure_to_pdf(plt, title):
        plt.figure(figsize=(10, 5))
        df = pd.DataFrame(summary_df, columns=['Project', 'Total Completed', 'Total Remaining'])
        df.plot(kind='bar', x='Project', stacked=True, color=['#4CAF50', '#FF9800'])

        plt.title(title)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()

        image_path = os.path.join(output_dir, f"bar_chart.png")
        plt.savefig(image_path, dpi=150)
        plt.close()

        pdf.add_page()
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(200, 10, txt=title.encode('latin-1', 'replace').decode('latin-1'), ln=True, align='C')
        pdf.image(image_path, x=10, y=30, w=190)  # Adjust x/y/w as needed


    # Add tables to PDF
    add_table_to_pdf(final_df, "Project Summary")
    add_figure_to_pdf(final_df, "Total Completed vs Total Remaining", "bar_chart.png")
    pdf.add_page()
    add_table_to_pdf(groupedC_df, "Detailed Summary by Type")

    output_file_pdf = os.path.join(output_dir, f"New_summary_output_{D}.pdf")
    pdf.output(output_file_pdf)


    print(f"\nSummary Excel file saved to: {output_file}")

# MAIN BLOCK
if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python graphgenerator.py <input_file> <output_dir>")
        sys.exit(1)

    input_file = sys.argv[1]
    output_dir = sys.argv[2]
    generate_summary(input_file, output_dir)
